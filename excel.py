"""openpyxl 기반 엑셀 내보내기."""
from __future__ import annotations

import os
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, numbers

import database as db


async def generate_excel(year: int, month: Optional[int] = None) -> str:
    """엑셀 파일 생성 후 경로 반환. month=None이면 연간."""
    wb = Workbook()
    months = [month] if month else list(range(1, 13))

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    income_font = Font(color="2E7D32")
    expense_font = Font(color="C62828")
    won_fmt = '#,##0'
    headers = ["날짜", "시간", "구분", "금액", "설명", "은행", "잔액", "카테고리"]

    for m in months:
        ws = wb.create_sheet(title=f"{year}-{m:02d}")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        txns = await db.get_transactions_for_range(
            f"{year}-{m:02d}-01", f"{year}-{m:02d}-31"
        )

        for i, tx in enumerate(txns, 2):
            ws.cell(row=i, column=1, value=tx["date"])
            ws.cell(row=i, column=2, value=tx["time"] or "")
            type_cell = ws.cell(row=i, column=3, value=tx["type"])
            type_cell.font = income_font if tx["type"] == "입금" else expense_font

            amt_cell = ws.cell(row=i, column=4, value=tx["amount"])
            amt_cell.number_format = won_fmt
            amt_cell.font = income_font if tx["type"] == "입금" else expense_font

            ws.cell(row=i, column=5, value=tx["description"] or "")
            ws.cell(row=i, column=6, value=tx["bank"] or "")

            if tx["balance"]:
                bal_cell = ws.cell(row=i, column=7, value=tx["balance"])
                bal_cell.number_format = won_fmt

            ws.cell(row=i, column=8, value=tx["category"] or "")

        # 합계 행
        if txns:
            r = len(txns) + 3
            ws.cell(row=r, column=3, value="입금 합계").font = Font(bold=True)
            ws.cell(row=r, column=4).value = f'=SUMIF(C2:C{r-2},"입금",D2:D{r-2})'
            ws.cell(row=r, column=4).number_format = won_fmt
            ws.cell(row=r, column=4).font = Font(bold=True, color="2E7D32")

            ws.cell(row=r+1, column=3, value="출금 합계").font = Font(bold=True)
            ws.cell(row=r+1, column=4).value = f'=SUMIF(C2:C{r-2},"출금",D2:D{r-2})'
            ws.cell(row=r+1, column=4).number_format = won_fmt
            ws.cell(row=r+1, column=4).font = Font(bold=True, color="C62828")

            ws.cell(row=r+2, column=3, value="순이익").font = Font(bold=True)
            ws.cell(row=r+2, column=4).value = f"=D{r}-D{r+1}"
            ws.cell(row=r+2, column=4).number_format = won_fmt
            ws.cell(row=r+2, column=4).font = Font(bold=True)

        # 열 너비
        widths = [12, 7, 8, 15, 20, 12, 15, 10]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = w

    # 기본 시트 제거
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    os.makedirs("./data", exist_ok=True)
    suffix = f"{month:02d}" if month else "all"
    path = f"./data/budget_{year}_{suffix}.xlsx"
    wb.save(path)
    return path
